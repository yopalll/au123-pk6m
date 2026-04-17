import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

"""
parse_body_category_excel.py
============================
Reads the 'body_bodyallcategory_uk.xlsx' scrape (Body/Category format)
and APPENDS clean JSON data to the existing VIYGO database JSON files.

Excel column mapping (1-based):
  Col 1  : Salon URL (treatwell.co.uk/place/...)
  Col 2  : Salon Name
  Col 3  : Rating (e.g. "5.0", "4.4")
  Col 4  : Review count (e.g. "5 reviews")
  Col 5  : Location (e.g. "Brooklands, Trafford")
  Col 6  : Service name (featured service)
  Col 7  : Service duration (e.g. "30 mins", "1 hr")
  Col 8  : Service price (e.g. "£90", "from")
  Col 9-55: Opening hours per day (day abbr, day full, open time, AM/PM, dash, close time, AM/PM)
  Col 56-63: Description sections
  Col 64,67,70: Additional prices
  Col 65,66,68,69: Additional service name/duration
  Col 71 : Off peak flag
  Col 72,73: Extended descriptions
  Col 74 : Brands info

Usage:
    python database/scripts/parse_body_category_excel.py

Output (appended/merged into):
    database/data/kota.json
    database/data/kategori.json
    database/data/salon.json
    database/data/service.json
    database/data/staff.json
    database/data/salon_images.json
"""

import openpyxl
import json
import re
import os
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).resolve().parent.parent   # database/
EXCEL_FILE = BASE_DIR / "body_bodyallcategory_uk.xlsx"
OUTPUT_DIR = BASE_DIR / "data"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Helpers ─────────────────────────────────────────────────────────────

def cell(ws, row, col):
    """Get cleaned cell value (1-based column index)."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def parse_rating(raw):
    """Extract numeric rating from string like '4.4'."""
    if not raw or raw in ("-.-", "0", "from"):
        return 0.0
    try:
        return round(float(raw), 2)
    except (ValueError, TypeError):
        return 0.0


def parse_reviews(raw):
    """Extract review count from '103 reviews' or '1 review'."""
    if not raw:
        return 0
    m = re.search(r'(\d+)\s*review', raw, re.IGNORECASE)
    return int(m.group(1)) if m else 0


def parse_price(raw):
    """Extract numeric price from '£90', '£35', 'from', etc."""
    if not raw:
        return None
    # Remove currency symbols, spaces
    cleaned = re.sub(r'[£€$,\s]', '', raw)
    try:
        return round(float(cleaned), 2)
    except (ValueError, TypeError):
        return None


def parse_duration(raw):
    """Convert '1 hr', '30 mins', '30 mins - 1 hr' → minutes (int)."""
    if not raw:
        return 60
    total = 0
    hr_m  = re.search(r'(\d+)\s*(?:hr|hrs|hour|hours)', raw, re.IGNORECASE)
    min_m = re.search(r'(\d+)\s*(?:min|mins|minutes)',  raw, re.IGNORECASE)
    if hr_m:
        total += int(hr_m.group(1)) * 60
    if min_m:
        total += int(min_m.group(1))
    return total if total > 0 else 60


def parse_location(raw):
    """
    Parse 'Brooklands, Trafford' or 'Old Ford, London' → city name.
    Takes the last meaningful part.
    """
    if not raw:
        return "Unknown"
    parts = [p.strip() for p in raw.split(",")]
    # Last part is usually the city/area
    return parts[-1].strip() if parts else "Unknown"


def parse_opening_hours_from_row(ws, row):
    """
    Parse opening hours from the complex multi-column structure.
    Columns 9-55 encode Mon-Sun open/close times.

    Pattern for each day (7 cols each, approx):
      day_abbr | day_full | open_H | AM/PM | dash | close_H | AM/PM

    We'll scan col 9 to 55 and look for AM/PM pairs.
    Strategy: collect all time+period strings and take first/last as open/close.
    """
    times = []
    for c in range(9, 56):
        v = cell(ws, row, c)
        # Look for time patterns like "10:00" and then next col "AM"/"PM"
        if re.match(r'^\d{1,2}:\d{2}$', v):
            period = cell(ws, row, c + 1)
            if period.upper() in ("AM", "PM"):
                h, m = map(int, v.split(":"))
                if period.upper() == "PM" and h != 12:
                    h += 12
                elif period.upper() == "AM" and h == 12:
                    h = 0
                times.append(h * 60 + m)

    if len(times) >= 2:
        open_m  = min(times)
        close_m = max(times)
        return (f"{open_m//60:02d}:{open_m%60:02d}",
                f"{close_m//60:02d}:{close_m%60:02d}")
    return ("09:00", "17:00")


def parse_description(ws, row):
    """Collect meaningful description from cols 56-63, 72-73."""
    parts = []
    for c in [56, 57, 58, 59, 60, 61, 62, 63, 72, 73]:
        v = cell(ws, row, c)
        if v and len(v) > 15:
            # Skip generic section headers
            if v.lower() not in ("nearest public transport:", "the team:",
                                  "nearest public transport", "the team",
                                  "what we like about the venue :",
                                  "what we like about the venue:"):
                parts.append(v)
    return " ".join(parts[:3]) if parts else None  # max 3 sections


def guess_category(service_name):
    """Map service name to VIYGO kategori."""
    n = service_name.lower()
    mapping = {
        "Hair":                ["hair", "haircut", "blowdry", "blow dry", "colour",
                                "color", "highlight", "balayage", "perm", "braid",
                                "extension", "trim", "shave", "barber", "keratin",
                                "scalp", "toner", "ombre", "root", "cornrow"],
        "Body":                ["body", "scrub", "wrap", "tan", "tanning",
                                "spray tan", "slimming", "exfoliation", "hifu",
                                "cellulite", "lymph", "detox", "contouring"],
        "Massage":             ["massage", "spa", "reflexology", "aromatherapy",
                                "deep tissue", "hot stone", "swedish", "thai",
                                "cupping", "shiatsu", "sports massage", "pregnancy massage"],
        "Face":                ["facial", "face", "dermaplaning", "microneedling",
                                "peel", "hydrating", "microdermabrasion", "gua sha",
                                "anti-aging", "anti-ageing"],
        "Nails":               ["nail", "manicure", "pedicure", "gel", "acrylic",
                                "shellac", "polish", "dipping"],
        "Hair Removal":        ["wax", "waxing", "threading", "laser", "hair removal",
                                "sugaring", "brazilian", "bikini", "epilat"],
        "Eyebrows & Lashes":   ["brow", "lash", "eyebrow", "eyelash", "tinting",
                                "lamination", "lash lift", "henna brow"],
        "Medical Aesthetics":  ["aesthetic", "filler", "dermal", "botox", "lip filler",
                                "micro", "chemical peel", "prp", "vitamin infusion"],
        "Counselling & Holistic": ["counsell", "holistic", "meditation", "reiki",
                                   "crystal", "chakra", "acupuncture", "hypnotherapy"],
    }
    for cat, keywords in mapping.items():
        for kw in keywords:
            if kw in n:
                return cat
    return "Body"  # default for this Body-category Excel


def load_json(filename):
    """Load existing JSON file, return list."""
    path = OUTPUT_DIR / filename
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_json(filename, data):
    """Save list as JSON file."""
    path = OUTPUT_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"   [OK] {filename}: {len(data)} records total")


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    print(f"[READ] Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(str(EXCEL_FILE), read_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1
    print(f"   Found {total_rows} data rows\n")

    # ── Load existing data ────────────────────────────────────────────
    print("[LOAD] Loading existing JSON files...")
    ex_kota     = load_json("kota.json")
    ex_kategori = load_json("kategori.json")
    ex_salon    = load_json("salon.json")
    ex_service  = load_json("service.json")
    ex_staff    = load_json("staff.json")
    ex_images   = load_json("salon_images.json")

    # Build lookup maps from existing data
    existing_salon_urls = {s.get("source_url", "") for s in ex_salon}
    city_id_map   = {k["nama_kota"]: k["id_kota"] for k in ex_kota}
    cat_id_map    = {c["name"]: c["id_kategori"] for c in ex_kategori}

    # ID counters (continue from where existing data left off)
    next_kota_id    = max((k["id_kota"]    for k in ex_kota),    default=0) + 1
    next_cat_id     = max((c["id_kategori"] for c in ex_kategori), default=0) + 1
    next_salon_id   = max((s["id_salon"]   for s in ex_salon),   default=0) + 1
    next_service_id = max((s["id_service"] for s in ex_service), default=0) + 1
    next_staff_id   = max((s["id_staff"]   for s in ex_staff),   default=0) + 1
    next_image_id   = max((i["id_salon_image"] for i in ex_images), default=0) + 1

    # ── Process rows ──────────────────────────────────────────────────
    new_salons   = []
    new_services = []
    new_staff    = []
    new_images   = []
    skipped      = 0
    processed    = 0

    for row in range(2, ws.max_row + 1):
        salon_url  = cell(ws, row, 1)
        salon_name = cell(ws, row, 2)

        if not salon_name:
            continue  # skip empty rows

        # Normalize URL (strip query params for dedup check)
        base_url = salon_url.split("?")[0] if salon_url else ""

        # Skip if already imported
        if base_url and base_url in existing_salon_urls:
            skipped += 1
            continue

        processed += 1
        salon_id = next_salon_id
        next_salon_id += 1

        # Track URL
        existing_salon_urls.add(base_url)

        # ── City / Kota ───────────────────────────────────────────────
        location_raw = cell(ws, row, 5)
        city_name    = parse_location(location_raw)
        if city_name not in city_id_map:
            city_id_map[city_name] = next_kota_id
            ex_kota.append({
                "id_kota":   next_kota_id,
                "nama_kota": city_name,
                "provinsi":  "England",
            })
            next_kota_id += 1
        id_kota = city_id_map[city_name]

        # ── Rating & Reviews ─────────────────────────────────────────
        rating       = parse_rating(cell(ws, row, 3))
        total_review = parse_reviews(cell(ws, row, 4))

        # ── Opening hours ─────────────────────────────────────────────
        open_time, close_time = parse_opening_hours_from_row(ws, row)

        # ── Description ───────────────────────────────────────────────
        description = parse_description(ws, row)
        if not description:
            description = f"{salon_name} is a professional beauty and wellness salon."

        # ── Salon record ──────────────────────────────────────────────
        new_salons.append({
            "id_salon":     salon_id,
            "id_user":      salon_id,
            "id_kota":      id_kota,
            "nama_salon":   salon_name,
            "alamat":       location_raw,
            "deskripsi":    description[:500],  # cap length
            "phone_number": None,
            "opening_time": open_time,
            "closing_time": close_time,
            "image_url":    None,
            "maps_url":     None,
            "latitude":     None,
            "longitude":    None,
            "rating":       rating,
            "total_review": total_review,
            "status":       "active",
            "source_url":   base_url,
        })

        # ── Services (primary + additional) ───────────────────────────
        # Primary service: cols 6, 7, 8
        services_raw = []
        svc_name_1 = cell(ws, row, 6)
        svc_dur_1  = cell(ws, row, 7)
        svc_price_1 = parse_price(cell(ws, row, 8))
        if svc_name_1:
            services_raw.append((svc_name_1, svc_dur_1, svc_price_1))

        # Additional services: cols (65,66,64), (68,69,67), (?, ?, 70)
        # Col 64 = price1, Col 65 = svc2 name, Col 66 = svc2 dur
        # Col 67 = price2, Col 68 = svc3 name, Col 69 = svc3 dur
        # Col 70 = price3
        for name_col, dur_col, price_col in [(65, 66, 64), (68, 69, 67)]:
            sn  = cell(ws, row, name_col)
            sd  = cell(ws, row, dur_col)
            sp  = parse_price(cell(ws, row, price_col))
            if sn and sn not in ("Off peak",):
                services_raw.append((sn, sd, sp))
        # Last price-only entry (col 70) attached to a service with no extra name
        extra_price = parse_price(cell(ws, row, 70))
        if extra_price and not services_raw:
            services_raw.append((svc_name_1 or "Body Treatment", "60 mins", extra_price))

        for svc_name, svc_dur, svc_price in services_raw:
            if not svc_name:
                continue
            cat_name   = guess_category(svc_name)
            # Ensure category exists
            if cat_name not in cat_id_map:
                slug = re.sub(r'[^a-z0-9]+', '-', cat_name.lower()).strip('-')
                cat_id_map[cat_name] = next_cat_id
                ex_kategori.append({
                    "id_kategori": next_cat_id,
                    "name":        cat_name,
                    "deskripsi":   f"Services related to {cat_name}",
                    "slug":        slug,
                    "icon_url":    None,
                    "is_active":   True,
                })
                next_cat_id += 1
            id_kategori = cat_id_map[cat_name]

            new_services.append({
                "id_service":  next_service_id,
                "id_salon":    salon_id,
                "id_kategori": id_kategori,
                "nama":        svc_name,
                "deskripsi":   None,
                "durasi":      parse_duration(svc_dur),
                "harga":       svc_price if svc_price else 0.0,
                "status":      "active",
            })
            next_service_id += 1

    # ── Merge & Save ─────────────────────────────────────────────────
    print(f"\n[MERGE] Processed {processed} new salons, skipped {skipped} duplicates\n")
    print("[WRITE] Saving merged JSON files...")

    save_json("kota.json",     ex_kota)
    save_json("kategori.json", ex_kategori)
    save_json("salon.json",    ex_salon + new_salons)
    save_json("service.json",  ex_service + new_services)
    save_json("staff.json",    ex_staff + new_staff)
    save_json("salon_images.json", ex_images + new_images)

    print(f"\n[SUMMARY]")
    print(f"   New salons added:   {len(new_salons)}")
    print(f"   New services added: {len(new_services)}")
    print(f"   Total kota:         {len(ex_kota)}")
    print(f"   Total kategori:     {len(ex_kategori)}")
    print(f"   Total salons:       {len(ex_salon) + len(new_salons)}")
    print(f"   Total services:     {len(ex_service) + len(new_services)}")
    print(f"   Output: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
