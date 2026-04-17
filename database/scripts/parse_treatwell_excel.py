import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

"""
parse_treatwell_excel.py
========================
Reads the messy Treatwell Excel scrape and produces clean JSON files
that map directly to the VIYGO database schema.

Usage:
    python database/scripts/parse_treatwell_excel.py

Output:
    database/data/kota.json
    database/data/kategori.json
    database/data/salon.json
    database/data/service.json
    database/data/staff.json
    database/data/salon_images.json
"""

import openpyxl
import json
import re
import os
from pathlib import Path

# ── Paths ─────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent  # database/
EXCEL_FILE = BASE_DIR / "treatwell-co-uk-2026-04-14.xlsx"
OUTPUT_DIR = BASE_DIR / "data"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Column index map (1-based, from inspection) ──────────────────────
COL = {
    "web_scraper_order":     1,
    "start_url":             2,
    "name_listing":          4,   # salon name from listing page
    "title":                56,   # salon name from detail page (more reliable)
    "rating_value":         59,   # e.g. "4.4"
    "address_full":         60,   # e.g. "191 Church Road, Yardley, Birmingham, B25 8UR"
    "address_locality_raw": 61,   # e.g. "Yardley,"
    "address_region":       98,   # e.g. "Birmingham"
    "postal_code_1":        87,
    "postal_code_2":       102,
    "opens":                62,   # e.g. "Open Today: 10:00 AM - 5:30 PM"
    "discount_info":        63,   # concatenated services+prices blob
    "review_criteria":      64,
    "review_criteria_2":    65,
    "total_reviews":        66,   # e.g. "103 reviews"
    "review_filter":        68,   # treatment‐level ratings
    "rating_5_stars":       69,
    "rating_4_stars":       70,
    "rating_3_stars":       71,
    "rating_1_star":        73,
    "team_member_service":  72,   # e.g. "HairFaceNailsHair removalMedical Aesthetics"
    "team_member":          79,   # e.g. "SolmazLead in hair stylist and aesthetics"
    "team_member_review":   80,
    "street_address":       81,
    "address_locality_2":   82,
    "item_page_link":       55,   # detail page URL
    "amenity":              96,
    "amenity_1":           101,
    "image_urls":          105,   # newline-separated CDN URLs
    "service_cat":          89,
    "service_cat_1":        90,
    "service_cat_2":        91,
    "service_cat_3":        92,
    "service_cat_5":        93,
    "service_cat_6":        94,

    # Listing-page data (less structured)
    "name2":                 5,   # service name from listing
    "data_duration":         6,   # e.g. "1 hr"
    "rating_listing":        7,   # rating from listing
    "data_location":         8,   # location from listing
    "data_price":           12,   # e.g. "£25"
    "description":          15,   # salon description
}


# ── Helpers ───────────────────────────────────────────────────────────

def cell(ws, row, col_key):
    """Get cleaned cell value."""
    val = ws.cell(row=row, column=COL[col_key]).value
    if val is None:
        return ""
    return str(val).strip()


def parse_rating(raw):
    """Extract numeric rating from string like '4.4' or '-.-'."""
    if not raw or raw == "-.-" or raw == "0":
        return 0.0
    try:
        return round(float(raw), 2)
    except (ValueError, TypeError):
        return 0.0


def parse_total_reviews(raw):
    """Extract review count from '103 reviews' or '1 review'."""
    if not raw:
        return 0
    match = re.search(r'(\d+)\s*review', raw)
    return int(match.group(1)) if match else 0


def parse_opening_hours(raw):
    """
    Parse 'Open Today: 10:00 AM - 5:30 PM' → ('10:00', '17:30')
    Returns 24h format.
    """
    if not raw:
        return ("09:00", "17:00")  # default

    match = re.search(r'(\d{1,2}:\d{2})\s*(AM|PM)\s*[-–]\s*(\d{1,2}:\d{2})\s*(AM|PM)', raw, re.IGNORECASE)
    if not match:
        return ("09:00", "17:00")

    def to_24h(time_str, period):
        h, m = map(int, time_str.split(":"))
        if period.upper() == "PM" and h != 12:
            h += 12
        elif period.upper() == "AM" and h == 12:
            h = 0
        return f"{h:02d}:{m:02d}"

    open_time = to_24h(match.group(1), match.group(2))
    close_time = to_24h(match.group(3), match.group(4))
    return (open_time, close_time)


def parse_address_city(address_full, locality_raw, region):
    """Extract city name from the address fields."""
    # Try region first (e.g. "Birmingham")
    if region and region not in ("", "None"):
        return region.strip().rstrip(",")

    # Try locality
    if locality_raw and locality_raw not in ("", "None"):
        return locality_raw.strip().rstrip(",")

    # Try to parse from full address
    if address_full:
        parts = [p.strip() for p in address_full.split(",")]
        if len(parts) >= 3:
            # Usually: street, area, city, postcode
            return parts[-2].strip()

    return "Unknown"


def parse_services_from_discount_info(raw):
    """
    Parse the concatenated Discount_Information blob into individual services.
    
    Example input:
    "Ladies - Haircuts & Blow Drys(5)from £30Ladies - Hair Treatments...(7)from £40...
     Children - Wash, Haircut & Blow Dry  45 minsShow Details£25Select..."
    
    Returns list of dicts: [{name, duration_minutes, price, category_hint}, ...]
    """
    if not raw:
        return []

    services = []

    # Pattern 1: "ServiceName  DurationShow Details£PriceSelect"
    # Pattern 2: "CategoryName(count)from £price"
    # Pattern 3: "ServiceName  DurationShow Details£PriceSelect"

    # First extract individual services (with Show Details pattern)
    service_pattern = re.compile(
        r'([A-Z][^£€$\d]*?)\s*'          # service name
        r'(\d+\s*(?:hr|hrs|min|mins|minutes|hour|hours)'
        r'(?:\s*\d+\s*(?:min|mins))?)\s*'  # duration
        r'(?:Show\s*Details)?'
        r'\s*[£€$]?\s*(\d+(?:\.\d{2})?)'  # price
        r'\s*(?:Select)?',
        re.IGNORECASE
    )

    for match in service_pattern.finditer(raw):
        name = match.group(1).strip()
        duration_raw = match.group(2).strip()
        price = float(match.group(3))

        # Clean up name
        name = re.sub(r'\s+', ' ', name)
        name = name.rstrip(' -')

        # Parse duration to minutes
        duration_min = parse_duration(duration_raw)

        if name and len(name) > 2:
            services.append({
                "nama": name,
                "durasi": duration_min,
                "harga": price,
            })

    # If we didn't find individual services, try category pattern
    if not services:
        cat_pattern = re.compile(
            r'([A-Z][^(£€$]*?)\s*\((\d+)\)\s*(?:from\s*)?[£€$]?\s*(\d+(?:\.\d{2})?)',
            re.IGNORECASE
        )
        for match in cat_pattern.finditer(raw):
            name = match.group(1).strip()
            price = float(match.group(3))
            if name and len(name) > 2:
                services.append({
                    "nama": name,
                    "durasi": 60,  # default 1hr
                    "harga": price,
                })

    return services


def parse_duration(raw):
    """Convert duration string to minutes. E.g. '1 hr 30 mins' → 90."""
    if not raw:
        return 60  # default

    total = 0
    hr_match = re.search(r'(\d+)\s*(?:hr|hrs|hour|hours)', raw, re.IGNORECASE)
    min_match = re.search(r'(\d+)\s*(?:min|mins|minutes)', raw, re.IGNORECASE)

    if hr_match:
        total += int(hr_match.group(1)) * 60
    if min_match:
        total += int(min_match.group(1))

    return total if total > 0 else 60


def parse_team_members(raw):
    """
    Parse team member names from the concatenated Team_Member field.
    E.g. "SolmazLead in hair stylist and aesthetics" → ["Solmaz"]
    E.g. "Mary" → ["Mary"]
    
    This is tricky because names are concatenated without separator.
    We'll use capital letter boundaries as heuristic.
    """
    if not raw:
        return []

    # Simple approach: just take the first word/name if it's short
    # Or split on common patterns
    raw = raw.strip()

    # Remove common suffixes
    suffixes = [
        r'Lead in.*', r'Senior.*', r'Junior.*', r'Owner.*',
        r'Manager.*', r'Stylist.*', r'Therapist.*', r'Specialist.*',
        r'Barber.*', r'Artist.*', r'Technician.*',
    ]
    cleaned = raw
    for suffix in suffixes:
        cleaned = re.sub(suffix, '', cleaned, flags=re.IGNORECASE).strip()

    # Split by capital letters (camelCase names)
    # e.g. "SolmazLindaJohn" → ["Solmaz", "Linda", "John"]
    names = re.findall(r'[A-Z][a-z]+', cleaned)

    if not names:
        # If no capital letter pattern, just use the whole string
        if len(cleaned) > 0 and len(cleaned) <= 30:
            names = [cleaned]

    # Deduplicate and limit
    seen = set()
    unique_names = []
    for n in names:
        if n.lower() not in seen and len(n) >= 2:
            seen.add(n.lower())
            unique_names.append(n)

    return unique_names[:5]  # max 5 staff per salon


def parse_image_urls(raw):
    """Parse newline-separated image URLs."""
    if not raw:
        return []

    urls = []
    for line in raw.split("\n"):
        url = line.strip()
        if url.startswith("http") and "treatwell.net" in url:
            urls.append(url)

    return urls


def guess_service_category(service_name):
    """Map a service name to one of our kategori based on keywords."""
    name_lower = service_name.lower()

    category_map = {
        "Hair": [
            "hair", "haircut", "blow dry", "blowdry", "colouring", "color",
            "highlight", "balayage", "perm", "cornrow", "braid", "twist",
            "extension", "restyle", "toner", "conditioning", "scalp",
            "keratin", "brazilian blow", "curly", "ombre", "root",
            "shampoo", "wash", "trim", "fade", "barber", "shave",
        ],
        "Face": [
            "facial", "face", "dermaplaning", "microneedling", "peel",
            "hydrating", "anti-aging", "botox",
        ],
        "Nails": [
            "nail", "manicure", "pedicure", "gel", "acrylic", "shellac",
            "polish",
        ],
        "Massage": [
            "massage", "spa", "reflexology", "aromatherapy", "deep tissue",
            "hot stone", "swedish", "thai",
        ],
        "Hair Removal": [
            "wax", "threading", "laser", "hair removal", "epilat",
            "sugaring", "brazilian", "bikini",
        ],
        "Body": [
            "body", "scrub", "wrap", "tan", "tanning", "spray tan",
            "slimming",
        ],
        "Medical Aesthetics": [
            "aesthetic", "filler", "dermal", "botox", "lip", "micro",
            "chemical peel",
        ],
        "Eyebrows & Lashes": [
            "brow", "lash", "eyebrow", "eyelash", "tinting", "lamination",
            "lash lift",
        ],
    }

    for category, keywords in category_map.items():
        for kw in keywords:
            if kw in name_lower:
                return category

    return "Other"


def parse_service_categories(ws, row):
    """Get all service categories from the Service_Category columns."""
    categories = set()
    for key in ["service_cat", "service_cat_1", "service_cat_2",
                "service_cat_3", "service_cat_5", "service_cat_6"]:
        val = cell(ws, row, key)
        if val and val != "All":
            categories.add(val.strip())

    # Also from team_member_service
    tms = cell(ws, row, "team_member_service")
    if tms:
        # Split concatenated categories like "HairFaceNailsHair removalMedical Aesthetics"
        known_cats = [
            "Medical Aesthetics", "Hair removal", "Hair", "Face",
            "Nails", "Massage", "Body", "Eyebrows & Lashes",
        ]
        remaining = tms
        for cat in known_cats:
            if cat in remaining:
                categories.add(cat)
                remaining = remaining.replace(cat, "", 1)

    return list(categories)


# ── Main Processing ──────────────────────────────────────────────────

def main():
    print(f"[READ] Reading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(str(EXCEL_FILE))
    ws = wb.active
    total_rows = ws.max_row - 1
    print(f"   Found {total_rows} data rows\n")

    # ── Collect unique cities & categories ────────────────────────────
    cities_set = set()
    categories_set = set()
    all_salons = []

    for row in range(2, ws.max_row + 1):
        title = cell(ws, row, "title")
        if not title:
            continue

        # City
        city = parse_address_city(
            cell(ws, row, "address_full"),
            cell(ws, row, "address_locality_raw"),
            cell(ws, row, "address_region"),
        )
        cities_set.add(city)

        # Categories from this salon
        cats = parse_service_categories(ws, row)
        categories_set.update([c.title() for c in cats])

        # Services
        services = parse_services_from_discount_info(cell(ws, row, "discount_info"))
        for svc in services:
            cat = guess_service_category(svc["nama"])
            categories_set.add(cat.title())

    # ── Build kota.json ──────────────────────────────────────────────
    kota_list = []
    city_id_map = {}
    for idx, city_name in enumerate(sorted(cities_set), start=1):
        city_id_map[city_name] = idx
        # Try to guess provinsi from the city
        provinsi = guess_provinsi(city_name)
        kota_list.append({
            "id_kota": idx,
            "nama_kota": city_name,
            "provinsi": provinsi,
        })

    # ── Build kategori.json ──────────────────────────────────────────
    kategori_list = []
    cat_id_map = {}
    for idx, cat_name in enumerate(sorted(categories_set), start=1):
        slug = re.sub(r'[^a-z0-9]+', '-', cat_name.lower()).strip('-')
        cat_id_map[cat_name] = idx
        kategori_list.append({
            "id_kategori": idx,
            "name": cat_name,
            "deskripsi": f"Services related to {cat_name}",
            "slug": slug,
            "icon_url": None,
            "is_active": True,
        })

    # Ensure 'Other' exists
    if "Other" not in cat_id_map:
        idx = len(kategori_list) + 1
        cat_id_map["Other"] = idx
        kategori_list.append({
            "id_kategori": idx,
            "name": "Other",
            "deskripsi": "Other services",
            "slug": "other",
            "icon_url": None,
            "is_active": True,
        })

    # ── Build salon, service, staff, salon_images ────────────────────
    salon_list = []
    service_list = []
    staff_list = []
    salon_images_list = []
    service_id = 0
    staff_id = 0
    image_id = 0

    for row in range(2, ws.max_row + 1):
        title = cell(ws, row, "title")
        if not title:
            continue

        salon_id = row - 1  # 1-based

        # City
        city = parse_address_city(
            cell(ws, row, "address_full"),
            cell(ws, row, "address_locality_raw"),
            cell(ws, row, "address_region"),
        )
        id_kota = city_id_map.get(city, 1)

        # Rating
        rating = parse_rating(cell(ws, row, "rating_value"))
        total_review = parse_total_reviews(cell(ws, row, "total_reviews"))

        # Opening hours
        open_time, close_time = parse_opening_hours(cell(ws, row, "opens"))

        # Address
        address = cell(ws, row, "address_full")

        # Description - from listing page description
        description = cell(ws, row, "description")
        if not description or len(description) < 10:
            description = f"{title} is a professional beauty and wellness salon."

        # Image (first one as main)
        images = parse_image_urls(cell(ws, row, "image_urls"))
        main_image = images[0] if images else None

        # Detail page link
        detail_link = cell(ws, row, "item_page_link")

        salon_list.append({
            "id_salon": salon_id,
            "id_user": salon_id,  # will be mapped to owner user
            "id_kota": id_kota,
            "nama_salon": title,
            "alamat": address,
            "deskripsi": description,
            "phone_number": None,
            "opening_time": open_time,
            "closing_time": close_time,
            "image_url": main_image,
            "maps_url": None,
            "latitude": None,
            "longitude": None,
            "rating": rating,
            "total_review": total_review,
            "status": "active",
            "source_url": detail_link,
        })

        # ── Services ─────────────────────────────────────────────────
        services = parse_services_from_discount_info(cell(ws, row, "discount_info"))
        for svc in services:
            service_id += 1
            cat_name = guess_service_category(svc["nama"]).title()
            id_kategori = cat_id_map.get(cat_name, cat_id_map.get("Other", 1))

            service_list.append({
                "id_service": service_id,
                "id_salon": salon_id,
                "id_kategori": id_kategori,
                "nama": svc["nama"],
                "deskripsi": None,
                "durasi": svc["durasi"],
                "harga": svc["harga"],
                "status": "active",
            })

        # ── Staff ────────────────────────────────────────────────────
        team_raw = cell(ws, row, "team_member")
        team_names = parse_team_members(team_raw)
        for name in team_names:
            staff_id += 1
            staff_list.append({
                "id_staff": staff_id,
                "id_salon": salon_id,
                "name": name,
                "profile_url": None,
                "status": "active",
            })

        # ── Salon Images ─────────────────────────────────────────────
        for img_idx, img_url in enumerate(images):
            image_id += 1
            salon_images_list.append({
                "id_salon_image": image_id,
                "id_salon": salon_id,
                "image_url": img_url,
                "is_primary": img_idx == 0,
                "urutan": img_idx + 1,
            })

    # ── Write JSON files ─────────────────────────────────────────────
    def save_json(filename, data):
        filepath = OUTPUT_DIR / filename
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"   [OK] {filename}: {len(data)} records")

    print("[WRITE] Writing JSON files...")
    save_json("kota.json", kota_list)
    save_json("kategori.json", kategori_list)
    save_json("salon.json", salon_list)
    save_json("service.json", service_list)
    save_json("staff.json", staff_list)
    save_json("salon_images.json", salon_images_list)

    print(f"\n[DONE] All files written to {OUTPUT_DIR}")
    print(f"\n[SUMMARY]")
    print(f"   Cities:       {len(kota_list)}")
    print(f"   Categories:   {len(kategori_list)}")
    print(f"   Salons:       {len(salon_list)}")
    print(f"   Services:     {len(service_list)}")
    print(f"   Staff:        {len(staff_list)}")
    print(f"   Images:       {len(salon_images_list)}")


def guess_provinsi(city):
    """Guess the region/county for UK cities."""
    uk_regions = {
        "Birmingham": "West Midlands",
        "London": "Greater London",
        "Manchester": "Greater Manchester",
        "Leeds": "West Yorkshire",
        "Liverpool": "Merseyside",
        "Bristol": "Avon",
        "Sheffield": "South Yorkshire",
        "Edinburgh": "Scotland",
        "Glasgow": "Scotland",
        "Cardiff": "Wales",
        "Belfast": "Northern Ireland",
        "Nottingham": "Nottinghamshire",
        "Leicester": "Leicestershire",
        "Newcastle": "Tyne and Wear",
        "Brighton": "East Sussex",
        "Oxford": "Oxfordshire",
        "Cambridge": "Cambridgeshire",
        "Bath": "Somerset",
        "York": "North Yorkshire",
        "Wolverhampton": "West Midlands",
        "Coventry": "West Midlands",
        "Dudley": "West Midlands",
        "Stourbridge": "West Midlands",
        "Solihull": "West Midlands",
        "Walsall": "West Midlands",
        "Sutton Coldfield": "West Midlands",
    }
    return uk_regions.get(city, "England")


if __name__ == "__main__":
    main()
